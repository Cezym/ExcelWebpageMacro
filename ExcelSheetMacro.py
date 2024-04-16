from openpyxl.reader.excel import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from DataTable import DataTable
from Interpreter import Interpreter


class ExcelSheetMacro(Interpreter):

    def __init__(self):
        super().__init__()
        self.data_table: DataTable = None
        self.path: str = None
        self.workbook: Workbook = None
        self.worksheet: Worksheet = None
        self.interpretations["open_sheet"] = self.open_sheet
        self.interpretations["get_cell"] = self.get_cell
        self.interpretations["save_cell"] = self.save_cell

    def get_cell(self, column, row):
        return self.data_table[column][row]

    def save_cell(self, column, row, value):
        if isinstance(column, str):
            column = self.data_table.labels[column]
        self.worksheet.cell(row+1, column+1, value)
        self.workbook.save(self.path)
        self.data_table.data[row][column] = value

    def open_sheet(self, path):
        self.path: str = path
        self.workbook: Workbook = load_workbook(filename=path)
        self.worksheet: Worksheet = self.workbook.active
        self.data_table = DataTable.fromExcel(path)


# macro = ExcelSheetMacro()
# inp: str = ""w
# while inp != "STOP":
#     #try:
#         inp = input()
#         print(macro.exec_line(inp))
#     #except Exception:
#     #    pass

if __name__ == "__main__":
    esm = ExcelSheetMacro()
    esm.open_sheet("Example.xlsx")
    print(esm.data_table)
    esm.exec_line("save_cell c 2 \"000\"")
    print(esm.data_table)
