from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from DataTable import DataTable
from ExcelSheetMacro import ExcelSheetMacro
from WebpageMacro import WebpageMacro


class WebpageExcelMacro(WebpageMacro, ExcelSheetMacro):
    def __init__(self):
        super().__init__()
        self.interpretations.update(ExcelSheetMacro().interpretations)
        self.interpretations["save_cell"] = self.save_cell
        self.interpretations["get_cell"] = self.get_cell
        self.interpretations["open_sheet"] = self.open_sheet

    def open_sheet(self, path):
        self.path: str = path
        self.workbook: Workbook = load_workbook(filename=path)
        self.worksheet: Worksheet = self.workbook.active
        self.data_table = DataTable.fromExcel(path)

    def save_cell(self, column, row, *value_parts):
        if isinstance(column, str):
            column = self.data_table.labels[column]

        print(value_parts)

        if len(value_parts) != 1:
            value = self.exec_parts(list(value_parts))
        else:
            value = value_parts[0]

        self.worksheet.cell(row + 1, column + 1, value)
        self.workbook.save(self.path)
        self.data_table.data[row][column] = value

    def get_cell(self, column, row):
        return self.data_table[column][row]

if __name__ == "__main__":
    wem = WebpageExcelMacro()
    print(wem.interpretations.keys())
    inp: str = ""
    while inp != "STOP":
        #try:
            inp = input()
            print(wem.exec_line(inp))
        #except Exception as e:
        #    print(e)